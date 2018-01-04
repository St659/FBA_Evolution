import openpyxl
import json
import re
import os


def metabolitesFromReactionString(rstring):
    """
    metabolitesFromReactionString(rstring)

    Abstract:
    Function to split a reaction string expression into a dictionaryof metabolit IDs
    and coefficients. Used when extracting FBA models from excel to json.

    Inputs:
    rstring - reaction string expression

    Outputs:
    a python dicionary

    """
    rcomponents = re.split('<==>|-->|<--', rstring)
    reaction_metabolites = dict()
    substrates = re.split('\+', rcomponents[0].replace(' ', ''))
    for s in substrates:
        if '*' in s:
            splits = re.split('\*', s)
            n = 0 - float(splits[0])  # negative as substrates are used up
            s = splits[1]
        else:
            n = -1
        reaction_metabolites[s] = n
    if (len(rcomponents) > 1):
        products = re.split('\+', rcomponents[1].replace(' ', ''))
        for s in products:
            if (s == ''):
                continue
            elif ('*' in s):
                splits = re.split('\*', s)
                n = float(splits[0])  # positive as products are produced
                s = splits[1]
            else:
                n = 1
            reaction_metabolites[s] = n
    return reaction_metabolites


def cobrajsonfromexcel(fname):
    """
    cobrajsonfromexcel(fname)

    Function to extract an FBA model from Excel (a la Jamie Wood format) and convert to
    JSON for use with cobrapy and escher. Requires package openpyxl json

    Inputs:
    fname - name and path to the excel file. Should be an .xlsx

    Outputs:
    None - the resulting json is written to a .json file

    """
    ##### open excel file
    #fullpath = os.path.abspath(fname)
    filename = fname
    #directory = os.path.dirname()
    model_id = '.'.join(filename.split('.')[:-1])
    print(model_id)
    print("Reading file: " + filename)
    try:
        wb = openpyxl.load_workbook(fname)
    except IOError as e:
        print('File could not be opened: {1}'.format(fname, e))
        return
    ##### check for all expected worksheets
    sheetNames = wb.get_sheet_names()
    expectedSheetNames = [
        'Compounds',
        'Reactions',
        'Biomass']
    missing = [x for x in expectedSheetNames if x not in sheetNames]
    if (len(missing) > 0):
        for m in missing:
            print('The following worksheets were missing from file: {0}'.format(', '.join(missing)))
        return
    ##### load universal bigg model
    bigg_fname = 'bigg_universal_model.json'
    try:
        universal = json.load(open(bigg_fname, 'r'))
    except IOError as e:
        print('Could not access universal BiGG model: {0}'.format(e))
        return
    universal_reactions = {k: v for (k, v) in zip([o['id'] for o in universal['reactions']], universal['reactions'])}
    universal_metabolites = {k: v for (k, v) in
                             zip([o['id'] for o in universal['metabolites']], universal['metabolites'])}
    ##### extract compound names from worksheet

    ws = wb['Compounds']
    original_list = list()
    i = 1
    while (ws['A' + str(i)].value != None):
        name = ws['A' + str(i)].value.strip()
        name = re.sub('\s', '_', name)  # remove whitespace
        if (name in original_list):
            print('Compound {0} is listed more than once'.format(name))
        original_list.append(name)
        i += 1
    if (len(set(original_list)) != len(original_list)):
        return 'Failure'
    ##### find all names that link to a metablite in bigg
    notfound = list()
    updated_metabolite_ids = dict()
    for name in original_list:
        if name in universal_metabolites:
            updated_metabolite_ids[name] = name
            continue
        # no match, create alternative versions of the name
        original_name = name[:]
        if (re.search('[\[\.,\s\(]', name)):
            name = re.sub('[\[\(]', '_', name)
            name = re.sub('[,\s\.\]\)]', '', name)
        nameArr = list(set([
            name,
            original_name,
            name + '_c',
            name.replace('-', '_'),
            name.replace('-', '__'),
            name.replace('-', '__') + '_c',
            name.replace('-', '_') + '_c',
        ]))
        matches = [True if n in universal_metabolites else False for n in nameArr]
        if (sum(matches) > 0):
            updated_metabolite_ids[original_name] = nameArr[matches.index(True)]
            continue
        # no matches for alt names in keys, search in bigg ids of each record
        for k in universal_metabolites:
            ur = universal_metabolites[k]
            bigg_ids = ur['notes']['original_bigg_ids']
            matches = [True if n in bigg_ids else False for n in nameArr]
            if (sum(matches) > 0):
                updated_metabolite_ids[original_name] = k
                break
        else:
            notfound.append(original_name)
    ##### print message if any metabolites were not found in bigg
    if (len(notfound) > 0):
        print("The following {0} metabolites were not found in bigg: {1}".format(len(notfound), ','.join(notfound)))
    ##### create an empty record for metabolites that dont match a bigg metabolite
    custom_metabolite = {
        'notes': {
            'original_bigg_ids': list()
        },
        'compartment': '',
        'id': '',
        'annotation': dict(),
        'name': ''
    }
    ##### make a dictionary for metabolites
    model_metabolites = dict()
    for name in original_list:
        original_name = name[:]
        if (name in updated_metabolite_ids):  # if true there is a bigg record
            name = updated_metabolite_ids[name]
            model_metabolites[original_name] = json.loads(json.dumps(universal_metabolites[name]))  # make a copy
            model_metabolites[original_name]['id'] = original_name
            model_metabolites[original_name]['bigg_id'] = name
        else:  # not in bigg
            model_metabolites[original_name] = json.loads(json.dumps(custom_metabolite))
            model_metabolites[original_name]['id'] = original_name
    ##### extract Reactions from worksheet
    ws = wb['Reactions']
    original_list = list()
    i = 1
    while (ws['A' + str(i)].value != None):
        name = ws['A' + str(i)].value.strip()
        name = re.sub('\s', '_', name)  # remove whitespace
        try:
            original_list.append({
                'id': name,
                'reaction': ws['B' + str(i)].value,
                'metabolites': metabolitesFromReactionString(ws['B' + str(i)].value),
                'lower_bound': float(ws['C' + str(i)].value),
                'upper_bound': float(ws['D' + str(i)].value),
                'subsystem': ws['F' + str(i)].value,
            })
        except ValueError:
            print(name + ' went wrong!')
        i += 1
    ##### remove all whitespace in metabolite ids
    for r in original_list:
        for m in r['metabolites']:
            if (re.search('\s', m)):
                newname = re.sub('\s', '_', m.strip())
                r['metabolites'][newname] = r['metabolites'][m]
                del r['metabolites'][m]
    ##### find all names that link to a reaction in bigg
    notfound = list()
    updated_reaction_ids = dict()
    for r in original_list:
        name = r['id']
        if name in universal_reactions:
            updated_reaction_ids[name] = name
            continue
        # no match, create alternative versions of the name
        original_name = name[:]
        if (re.search('[\[\.,\s\(]', name)):
            name = re.sub('[\[\(]', '_', name)
            name = re.sub('[,\s\.\]\)]', '', name)
        nameArr = list(set([
            name,
            original_name,
            name.replace('-', '_'),
            name.replace('-', '__')
        ]))
        matches = [True if n in universal_reactions else False for n in nameArr]
        if (sum(matches) > 0):
            updated_reaction_ids[original_name] = nameArr[matches.index(True)]
            continue
        # no matches for alt names in keys, search in bigg ids of each record
        for k in universal_reactions:
            ur = universal_reactions[k]
            bigg_ids = ur['notes']['original_bigg_ids']
            matches = [True if n in bigg_ids else False for n in nameArr]
            if (sum(matches) > 0):
                updated_reaction_ids[original_name] = k
                break
        else:
            notfound.append(original_name)
    ##### print message if any metabolites were not found in bigg
    if (len(notfound) > 0):
        print("The following {0} reactions were not found in bigg: {1}".format(len(notfound), ','.join(notfound)))
    ##### create an empty reaction record to use if any were not found in bigg
    custom_reaction = {
        "name": "",
        "upper_bound": 999999,
        "lower_bound": -999999,
        "notes": {
            "original_bigg_ids": list()
        },
        "annotation": dict(),
        "metabolites": dict(),
        "id": "",
        "gene_reaction_rule": ""
    }
    ##### make a dictionary for reactions
    model_reactions = dict()
    for r in original_list:
        original_name = r['id']
        name = original_name[:]
        if (name in updated_reaction_ids):  # if true there is a bigg record
            name = updated_reaction_ids[name]
            model_reactions[original_name] = json.loads(json.dumps(universal_reactions[name]))
            model_reactions[original_name]['id'] = original_name
            model_reactions[original_name]['bigg_id'] = name
            model_reactions[original_name]['metabolites'] = r['metabolites']
            model_reactions[original_name]['upper_bound'] = r['upper_bound']
            model_reactions[original_name]['lower_bound'] = r['lower_bound']
        else:  # not in bigg
            model_reactions[original_name] = json.loads(json.dumps(custom_reaction))
            model_reactions[original_name]['id'] = original_name
            model_reactions[original_name]['metabolites'] = r['metabolites']
            model_reactions[original_name]['upper_bound'] = r['upper_bound']
            model_reactions[original_name]['lower_bound'] = r['lower_bound']
    ##### load biomass/objective reaction
    ws = wb['Biomass']
    original_list = dict()
    i = 1
    while (ws['A' + str(i)].value != None):
        name = ws['A' + str(i)].value
        number = float(ws['B' + str(i)].value)
        include = int(ws['C' + str(i)].value)
        if (include > 0):
            original_list[name] = number
        i += 1
    ##### add it to the model as a reaction
    model_reactions['Biomass'] = json.loads(json.dumps(custom_reaction))
    model_reactions['Biomass']['id'] = 'Biomass'
    model_reactions['Biomass']['metabolites'] = original_list
    model_reactions['Biomass']['objective_coefficient'] = 1  # set it as the objective
    ##### check that all reaction metabolites match up with a metabolite key
    missing = list()
    for r in model_reactions:
        reaction = model_reactions[r]
        for m in reaction['metabolites']:
            if m not in model_metabolites:
                missing.append(m)
    ##### print message if any were missing
    if (len(missing) > 0):
        print("The following {0} reaction metabolites were not found in the metabolite list: {1}".format(len(missing),
                                                                                                         ','.join(
                                                                                                             missing)))


    ##### build the model obj
    model = dict()
    model_id = '.'.join(fname.split('/')[-1].split('.')[:-1])
    print('Model id ' + model_id)
    model['id'] = model_id
    model['metabolites'] = [model_metabolites[k] for k in model_metabolites]
    model['reactions'] = [model_reactions[k] for k in model_reactions]
    model['genes'] = list()
    ##### write to file
    outfname = os.path.join( model_id + '.json')
    with open(outfname, 'w') as f:
        f.write(json.dumps(model))
    print("Model output to: {0}".format(model_id + '.json'))
    return model_id + '.json'

